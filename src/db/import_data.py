import sqlite3
import openpyxl
import random
import datetime
from faker import Faker

wb = openpyxl.load_workbook('NBA_Dataset.xlsx')

conn = sqlite3.connect("nba.db")
cur = conn.cursor()

#teams

ws = wb['teams']

teams = []
team_map = {}

for row in ws.iter_rows():
    row_vals = []
    for cell in row:
        row_vals.append(cell.value)
    team_map[row_vals[1]] = row_vals[0]
    teams.append(row_vals)
    
cur.execute("DELETE FROM team;")
cur.executemany("""
        INSERT INTO team (team_id, team_name_short, team_name)
        VALUES (?, ?, ?);
    """, teams)

conn.commit()

#players

ws = wb['players']

players = []
player_map = {}

for row in ws.iter_rows():
    row_vals = []
    for cell in range(0, len(row)):
        if cell != 2:
            row_vals.append(row[cell].value)
        else:
            row_vals.append(team_map[row[cell].value])
    players.append(row_vals)
    player_map[row_vals[1]] = row_vals[0]

cur.execute("DELETE FROM player;")
cur.executemany("""
        INSERT INTO player (player_id, player_name, team_id, age, height, weight, college, birth_country, draft_year, draft_round)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
    """, players)

conn.commit()

#contracts

ws = wb['contracts']

contracts = []
values = [100000, 240000, 1300000, 400000, 600000, 10000000]

for row in ws.iter_rows():
    row_vals = []
    if not row[0].value in player_map:
        continue

    for cell in range(0, len(row)):
        if cell == 0:
            row_vals.append(player_map[row[cell].value])
        elif cell == 1:
            row_vals.append(team_map[row[cell].value])
        else:
            year = str(row[cell].value).split('-')
            row_vals.append(year[0])
            if int(year[1]) >= 0 and int(year[1]) < 96:
                row_vals.append("20" + year[1])
            else:
                row_vals.append("19" + year[1])
    row_vals.append(random.choice(values))
    contracts.append(row_vals)

cur.execute("DELETE FROM contract;")
cur.executemany("""
        INSERT INTO contract (player_id, team_id, start_date, end_date, total_value)
        VALUES (?, ?, ?, ?, ?);
    """, contracts)

conn.commit()

#match
matches = []
fake = Faker()
start_date = datetime.date(year=2021, month=1, day=1)
end_date = datetime.date(year=2021, month=11, day=30)

while len(matches) < 20:
    opt1 = random.choice(teams)
    opt2 = random.choice(teams)
    if opt1[0] != opt2[0]:
        date = fake.date_between(start_date, end_date)
        win_prob = [1, 2, 3, 4, 5, 6]
        matches.append([opt1[0], opt2[0], date.strftime("%m/%d/%Y"), opt1[0] if random.choice(win_prob) % 2 == 0 else opt2[0]])

cur.execute("DELETE FROM match;")
cur.executemany("""
        INSERT INTO match (team_id_1, team_id_2, date, winning_team) VALUES (?, ?, ?, ?);
    """, matches)

conn.commit()

#sponsor

companies = ['Mountain Dew', 'Red Bull', 'EA', 'Microsoft', 'Nike', 'Adidas', 'Reebok', 'Tesla', 'Vivint', 'Ford', 'Dodge', 'Chevrolet']

sponsors = []

for team in teams:
    sponsors.append([team[0], random.choice(companies)])

cur.execute("DELETE FROM sponsor;")
cur.executemany("""
        INSERT INTO sponsor (team_id, sponsor_name) VALUES (?, ?);
    """, sponsors)

conn.commit()

#owner

owners = [
    [1, 'Jerry Reinsdorf'],
    [2, 'Steve Ballmer'],
    [3, 'Maple Leaf Sports & Entertainment'],
    [4, 'Mark Cuban'],
    [5, 'Micky Arison'],
    [6, 'Tilman Fertitta'],
    [7, 'Jeanie Buss'],
    [8, 'Tony Ressler'],
    [9, 'Marc Lasry and Wes Edens'],
    [10, 'Stan Kroenke'],
    [11, ''],
    [12, 'Jody Allen'],
    [13, ''],
    [14, 'Joe Tsai'],
    [15, 'Wyc Grousbeck'],
    [16, 'Herb Simon'],
    [17, 'Vivek Ranadivé'],
    [18, 'Glen Taylor'],
    [19, 'Joshua Harris'],
    [20, 'DeVos family'],
    [21, 'Julianna Hawn Holt'],
    [22, 'Robert Sarver'],
    [23, 'Tom Gores'],
    [24, 'Michael Jordan'],
    [25, 'Dan Gilbert'],
    [26, 'Joe Lacob and Peter Guber'],
    [27, 'Ryan Smith'],
    [28, 'Ted Leonsis'],
    [29, 'James Dolan'],
    [30, 'Robert J. Pera'],
    [31, ''],
    [32, ''],
    [33, ''],
    [34, 'Clayton I. Bennett'],
    [35, ''],
    [36, 'Gayle Benson']
]

cur.execute("DELETE FROM owner;")
cur.executemany("""
        INSERT INTO owner (team_id, owner_name) VALUES (?, ?);
    """, owners)

conn.commit()

conn.close()
